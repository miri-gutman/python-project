import openpyxl as xl


def num_sheets(path):
    wb=xl.load_workbook(path)
    num_sheets = len(wb.sheetnames)
    return num_sheets

def sum_column(file_path, sheet_name, column_letter):
    wb = xl.load_workbook(file_path)
    sheet = wb[sheet_name]

    column_values = [sheet[column_letter + str(row)].value for row in range(2, sheet.max_row + 1) if
                     sheet[column_letter + str(row)].value is not None]
    column_sum = sum(column_values)

    return column_sum

def average_column(file_path, sheet_name, column_letter):
        wb = xl.load_workbook(file_path)
        sheet = wb[sheet_name]

        column_values = [sheet[column_letter + str(row)].value for row in range(1, sheet.max_row + 1) if
                         sheet[column_letter + str(row)].value is not None]

        if not column_values:
            return 0  # Handle the case where the column is empty

        column_average = sum(column_values) / len(column_values)

        return column_average

