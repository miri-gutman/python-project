import openpyxl as xl
import matplotlib.pyplot as plt
import os

def num_sheets(path):
    wb=xl.load_workbook(path)
    num_sheets = len(wb.sheetnames)
    return num_sheets

def num_fields(path):
    total_sum = 0
    workbook = xl.load_workbook(path)
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        for row in current_sheet.iter_rows(values_only=True):
           for cell in row:
               if cell is not None:
                total_sum += cell

    workbook.close()
    return total_sum

def sum_column_graph(path):
    data={}
    workbook = xl.load_workbook(path)
    for sheet in workbook.sheetnames:
        current_sheet = workbook[sheet]
        total_sum = 0
        for row in current_sheet.iter_rows(values_only=True):
            for cell in row:
                if cell is not None:
                    total_sum += cell
        data[sheet]= total_sum

    workbook.close()
    pages=list(data.keys())
    sum=list(data.values())
    plt.bar(pages,sum,color='red',width=0.4)
    plt.xlabel("pages")
    plt.ylabel("sum of cells in the page")
    plt.title("how many sum of cells in the page")
    plt.show()
def avg_column_graph(list_files):
    data={}
    for file in list_files:
        sum = num_fields(file)
        num_pages = num_sheets(file)
        data[file]=sum/num_pages

    pages=list(data.keys())
    avg=list(data.values())
    plt.bar(pages,avg,color='red',width=0.4)
    plt.xlabel("average")
    plt.title("average sum to file")
    plt.show()

# sum_column_graph(os.path.join(os.getcwd(),'new.xlsx'))
# list_file=[os.path.join(os.getcwd(),'new.xlsx'),os.path.join(os.getcwd(),'excel_file.xlsx')]
# avg_column_graph(list_file)